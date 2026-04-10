import importlib.util
import pkgutil

if not hasattr(pkgutil, 'get_loader'):
    def get_loader(name):
        spec = importlib.util.find_spec(name)
        return None if spec is None else spec.loader
    pkgutil.get_loader = get_loader

from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, send_file, send_from_directory
from functools import wraps
from sqlalchemy import text
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
import io
import csv
import json
import math
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'fallback-dev-key')
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATABASE_PATH = os.path.join(BASE_DIR, 'instance', 'construction_crm.db')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL') or f'sqlite:///{DATABASE_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Allowed file extensions for upload
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def handle_errors(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            app.logger.error(f"Error in {f.__name__}: {str(e)}")
            return jsonify({'error': str(e)}), 500
    return wrapper


def is_na(val):
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == '':
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return False


def clean_value(val):
    if is_na(val):
        return ''
    return str(val).strip()


def read_spreadsheet(file, sheet_name=None):
    filename = file.filename.lower()
    if filename.endswith('.csv'):
        payload = file.read().decode('utf-8-sig')
        file.seek(0)
        reader = csv.DictReader(io.StringIO(payload))
        return [row for row in reader]

    if not filename.endswith(('.xlsx', '.xls')):
        raise ValueError('Unsupported file type')

    file.seek(0)
    wb = load_workbook(file, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return data


def write_to_excel_buffer(data, sheet_name='Sheet1'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for item in data:
            ws.append([item.get(k, '') for k in headers])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Database Models
project_contacts = db.Table('project_contacts',
    db.Column('project_id', db.Integer, db.ForeignKey('project.id'), primary_key=True),
    db.Column('contact_id', db.Integer, db.ForeignKey('contact.id'), primary_key=True)
)

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    avatar = db.Column(db.String(200), default='default.png')
    last_login = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    role = db.Column(db.String(20), default='staff')
    is_active = db.Column(db.Boolean, default=True)
    two_factor_enabled = db.Column(db.Boolean, default=False)
    notification_email = db.Column(db.Boolean, default=True)
    notification_login_alert = db.Column(db.Boolean, default=True)
    last_login_ip = db.Column(db.String(45))
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Contact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    company = db.Column(db.String(100))
    type = db.Column(db.String(50))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def relation(self):
        """Parse relation from notes (relation:good|average|bad)"""
        import re
        if not self.notes:
            return None
        m = re.search(r'relation:(good|average|bad)', self.notes)
        return m.group(1) if m else None

class Provider(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact_person = db.Column(db.String(100))
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    address = db.Column(db.String(200))
    service_type = db.Column(db.String(100))
    notes = db.Column(db.Text)

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    client_name = db.Column(db.String(100))
    address = db.Column(db.String(200))
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    status = db.Column(db.String(50))
    description = db.Column(db.Text)
    budget_total = db.Column(db.Float, default=0.0)
    selling_price = db.Column(db.Float, default=0.0)
    client_receipts = db.Column(db.Float, default=0.0)
    
    # Relationships
    contacts = db.relationship('Contact', secondary=project_contacts, backref=db.backref('projects', lazy='dynamic'))

class Tool(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50))
    quantity = db.Column(db.Integer, default=1)
    purchase_price = db.Column(db.Float)
    current_value = db.Column(db.Float)
    condition = db.Column(db.String(20), default='Good')  # New, Good, Used, Damaged
    location = db.Column(db.String(100))
    purchase_date = db.Column(db.Date)
    notes = db.Column(db.Text)
    
    def __repr__(self):
        return f'<Tool {self.name}>'

class ProjectWorker(db.Model):
    """Model for tracking workers (subcontractors and daily workers) assigned to projects"""
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id', ondelete='CASCADE'), nullable=False)
    contact_id = db.Column(db.Integer, db.ForeignKey('contact.id'), nullable=False)
    worker_type = db.Column(db.String(50), nullable=False)  # 'subcontractor' or 'daily_worker'
    role = db.Column(db.String(200))  # Role/task for this project
    start_date = db.Column(db.Date)
    contract_amount = db.Column(db.Float, default=0.0)  # For subcontractors
    daily_rate = db.Column(db.Float, default=0.0)  # For daily workers
    days_worked = db.Column(db.Float, default=0.0)  # Total days worked for daily workers
    amount_paid = db.Column(db.Float, default=0.0)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    contact = db.relationship('Contact', backref='project_assignments')
    project = db.relationship('Project', backref=db.backref('workers', lazy=True, cascade='all, delete-orphan', passive_deletes=True))
    attendances = db.relationship('Attendance', backref='project_worker', lazy=True, cascade='all, delete-orphan', passive_deletes=True)
    payments = db.relationship('Payment', backref='project_worker', lazy=True, cascade='all, delete-orphan', passive_deletes=True)
    
    @property
    def remaining_balance(self):
        if self.worker_type == 'subcontractor':
            return self.contract_amount - self.amount_paid
        return 0.0
    
    @property
    def status(self):
        if self.worker_type == 'subcontractor':
            if self.amount_paid >= self.contract_amount:
                return 'paid'
            elif self.amount_paid > 0:
                return 'partial'
            return 'pending'
        return 'active'

    @property
    def calculated_days_worked(self):
        """Calculate total days worked from attendances"""
        try:
            return sum([a.days or 0 for a in self.attendances])
        except Exception:
            return 0

class ProjectExpense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    date = db.Column(db.Date)
    nature = db.Column(db.String(200))
    amount = db.Column(db.Float)
    comment = db.Column(db.Text)
    category = db.Column(db.String(50))
    
    project = db.relationship('Project', backref=db.backref('expenses', lazy=True))

class ProjectReceipt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    date = db.Column(db.Date)
    amount = db.Column(db.Float)
    comment = db.Column(db.Text)
    
    project = db.relationship('Project', backref=db.backref('receipts', lazy=True))

class ProjectFinancialParams(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), unique=True)
    sale_price = db.Column(db.Float, default=0)
    estimated_budget = db.Column(db.Float, default=0)
    devis_ref = db.Column(db.Float, default=0)
    
    project = db.relationship('Project', backref=db.backref('financial_params', uselist=False))

class ProjectMemo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    date = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    project = db.relationship('Project', backref=db.backref('memos', lazy=True))

class ProjectPlan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    date = db.Column(db.Date, nullable=False)
    file_path = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    project = db.relationship('Project', backref=db.backref('plans', lazy=True))

class ProjectContract(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    date = db.Column(db.Date, nullable=False)
    file_path = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    project = db.relationship('Project', backref=db.backref('contracts', lazy=True))

class ProjectInvoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    date = db.Column(db.Date, nullable=False)
    file_path = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    project = db.relationship('Project', backref=db.backref('invoices', lazy=True))


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_worker_id = db.Column(db.Integer, db.ForeignKey('project_worker.id', ondelete='CASCADE'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    days = db.Column(db.Float, default=0.0)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Attendance pw={self.project_worker_id} date={self.date} days={self.days}>'


class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_worker_id = db.Column(db.Integer, db.ForeignKey('project_worker.id', ondelete='CASCADE'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Float, default=0.0)
    method = db.Column(db.String(50))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Payment pw={self.project_worker_id} date={self.date} amount={self.amount}>'

class ProjectTask(db.Model):
    """Model for project tasks in Gantt chart"""
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    progress = db.Column(db.Integer, default=0)
    parent_id = db.Column(db.Integer, db.ForeignKey('project_task.id'), nullable=True)
    dependency_type = db.Column(db.String(10))  # 'FS', 'SS', 'FF', 'SF' or NULL
    assigned_to = db.Column(db.String(200))
    status = db.Column(db.String(20), default='pending')  # 'pending', 'in_progress', 'completed'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    project = db.relationship('Project', backref=db.backref('tasks', lazy=True, cascade='all, delete-orphan'))
    parent = db.relationship('ProjectTask', remote_side=[id], backref=db.backref('children', lazy=True))

    def __repr__(self):
        return f'<ProjectTask {self.name} project={self.project_id}>'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ==================== AUTH ROUTES ====================
@app.route('/project/<int:project_id>/financial/update-summary', methods=['POST'])
@login_required
def update_financial_summary(project_id):
    """Update financial summary - syncs with Project (selling_price, client_receipts) and ProjectFinancialParams"""
    try:
        project = Project.query.get_or_404(project_id)
        financial = ProjectFinancialParams.query.filter_by(project_id=project_id).first()
        if not financial:
            financial = ProjectFinancialParams(project_id=project_id)
            db.session.add(financial)
        
        data = request.get_json()
        
        # Update Project (same fields as project detail page)
        if 'sale_price' in data:
            project.selling_price = float(data.get('sale_price', 0))
        if 'client_received' in data:
            project.client_receipts = float(data.get('client_received', 0))
        
        # Update ProjectFinancialParams
        financial.sale_price = float(data.get('sale_price', project.selling_price or 0))
        financial.estimated_budget = float(data.get('estimated_budget', financial.estimated_budget or 0))
        financial.devis_ref = float(data.get('devis_ref', financial.devis_ref or 0))
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/')
def index():
    if current_user.is_authenticated:
        # Get project statistics
        active_projects = Project.query.filter_by(status='in_progress').count()
        completed_projects = Project.query.filter_by(status='completed').count()
        planned_projects = Project.query.filter_by(status='planned').count()
        
        return render_template('index.html', 
                             active_projects=active_projects,
                             completed_projects=completed_projects,
                             planned_projects=planned_projects)
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username').lower().strip()
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            # Update last login time and IP
            user.last_login = datetime.utcnow()
            user.last_login_ip = request.remote_addr
            db.session.commit()
            login_user(user)
            return redirect(url_for('index'))
        flash('Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if not current_user.check_password(current_password):
            flash('Current password is incorrect')
        elif new_password != confirm_password:
            flash('New passwords do not match')
        elif len(new_password) < 6:
            flash('Password must be at least 6 characters')
        else:
            current_user.set_password(new_password)
            db.session.commit()
            flash('Password changed successfully!')
            return redirect(url_for('index'))
    
    return render_template('change_password.html')

@app.route('/account-management')
@login_required
def account_management():
    # Mock current session data (replace with actual session tracking)
    current_session = {
        'browser': 'Chrome',
        'os': 'Windows'
    }
    
    # Mock active sessions (replace with actual session tracking)
    active_sessions = []
    
    return render_template('account_management.html', 
                          current_session=current_session,
                          active_sessions=active_sessions)

@app.route('/users')
@login_required
def list_users():
    """Display all users with their IDs and login info"""
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/account/upload-avatar', methods=['POST'])
@login_required
def upload_avatar():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type'})
    
    # Create uploads directory if it doesn't exist
    upload_dir = os.path.join(app.static_folder, 'uploads', 'avatars')
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generate unique filename
    filename = secure_filename(f"{current_user.id}_{int(datetime.utcnow().timestamp())}_{file.filename}")
    filepath = os.path.join(upload_dir, filename)
    
    # Save file
    file.save(filepath)
    
    # Update user avatar
    current_user.avatar = filename
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/account/update-email', methods=['POST'])
@login_required
def update_email():
    data = request.get_json()
    new_email = data.get('email')
    
    if not new_email:
        return jsonify({'success': False, 'error': 'Email is required'})
    
    # Check if email already exists
    existing_user = User.query.filter_by(email=new_email).first()
    if existing_user and existing_user.id != current_user.id:
        return jsonify({'success': False, 'error': 'Email already in use'})
    
    current_user.email = new_email
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/account/update-notifications', methods=['POST'])
@login_required
def update_notifications():
    data = request.get_json()
    notification_type = data.get('type')
    enabled = data.get('enabled', False)
    
    if notification_type == 'email':
        current_user.notification_email = enabled
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/account/update-login-alerts', methods=['POST'])
@login_required
def update_login_alerts():
    data = request.get_json()
    enabled = data.get('enabled', False)
    
    current_user.notification_login_alert = enabled
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/account/revoke-session', methods=['POST'])
@login_required
def revoke_session():
    # Implement session revocation logic
    return jsonify({'success': True, 'message': 'Session revoked'})

@app.route('/account/sign-out-all', methods=['POST'])
@login_required
def sign_out_all_devices():
    # Implement sign out all devices logic
    return jsonify({'success': True, 'message': 'Signed out from all other devices'})

@app.route('/account/delete', methods=['POST'])
@login_required
def delete_account():
    # Prevent deletion of admin accounts
    if current_user.role == 'admin':
        return jsonify({'success': False, 'error': 'Cannot delete admin account'})
    
    # Delete user account
    db.session.delete(current_user)
    db.session.commit()
    
    return jsonify({'success': True})

# ==================== CONTACT ROUTES ====================
@app.route('/contacts')
@login_required
def contacts():
    contacts_list = Contact.query.all()
    return render_template('contacts.html', contacts=contacts_list)

@app.route('/contacts/add', methods=['POST'])
@login_required
def add_contact():
    contact = Contact(
        name=request.form.get('name'),
        email=request.form.get('email'),
        phone=request.form.get('phone'),
        company=request.form.get('company'),
        type=request.form.get('type'),
        notes=request.form.get('notes')
    )
    db.session.add(contact)
    db.session.commit()
    flash('Contact added successfully')
    return redirect(url_for('contacts'))

@app.route('/contacts/delete/<int:id>')
@login_required
def delete_contact(id):
    contact = Contact.query.get_or_404(id)
    db.session.delete(contact)
    db.session.commit()
    flash('Contact deleted')
    return redirect(url_for('contacts'))

@app.route('/contacts/<int:contact_id>/details')
@login_required
def get_contact_details(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    return jsonify({
        'id': contact.id,
        'name': contact.name,
        'email': contact.email,
        'phone': contact.phone,
        'company': contact.company,
        'type': contact.type,
        'notes': contact.notes
    })

@app.route('/contacts/<int:contact_id>/edit', methods=['POST'])
@login_required
def edit_contact(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    contact.name = request.form.get('name')
    contact.email = request.form.get('email')
    contact.phone = request.form.get('phone')
    contact.company = request.form.get('company')
    contact.type = request.form.get('type')
    contact.notes = request.form.get('notes')
    
    db.session.commit()
    flash('Contact updated successfully')
    return redirect(url_for('contacts'))

@app.route('/contacts/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_contacts():
    data = request.get_json()
    ids = data.get('ids', [])
    
    if ids:
        Contact.query.filter(Contact.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'success': True, 'count': len(ids)})
    
    return jsonify({'success': False}), 400

@app.route('/contacts/<int:contact_id>/update-type', methods=['POST'])
@login_required
def update_contact_type(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    data = request.get_json()
    new_type = data.get('type')
    if new_type:
        contact.type = new_type
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 400

@app.route('/contacts/<int:contact_id>/update-relation', methods=['POST'])
@login_required
def update_contact_relation(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    data = request.get_json()
    relation = data.get('relation', '')
    
    notes = contact.notes or ''
    import re
    notes = re.sub(r'relation:(good|average|bad)\s*', '', notes)
    
    if relation:
        notes = notes.strip() + f' relation:{relation}'
    
    contact.notes = notes.strip()
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/contacts')
@login_required
def api_contacts():
    contacts = Contact.query.all()
    return jsonify([{
        'id': contact.id,
        'name': contact.name,
        'phone': contact.phone,
        'email': contact.email,
        'company': contact.company,
        'type': contact.type
    } for contact in contacts])

# ==================== PROVIDER ROUTES ====================
@app.route('/providers')
@login_required
def providers():
    providers_list = Provider.query.all()
    return render_template('providers.html', providers=providers_list)

@app.route('/providers/add', methods=['POST'])
@login_required
def add_provider():
    provider = Provider(
        name=request.form.get('name'),
        contact_person=request.form.get('contact_person'),
        email=request.form.get('email'),
        phone=request.form.get('phone'),
        address=request.form.get('address'),
        service_type=request.form.get('service_type'),
        notes=request.form.get('notes')
    )
    db.session.add(provider)
    db.session.commit()
    flash('Provider added successfully')
    return redirect(url_for('providers'))

@app.route('/providers/delete/<int:id>')
@login_required
def delete_provider(id):
    provider = Provider.query.get_or_404(id)
    db.session.delete(provider)
    db.session.commit()
    flash('Provider deleted')
    return redirect(url_for('providers'))

# ==================== PROJECT ROUTES ====================
@app.route('/projects')
@login_required
def projects():
    projects_list = Project.query.all()
    return render_template('projects.html', projects=projects_list)

@app.route('/projects/add', methods=['POST'])
@login_required
def add_project():
    project = Project(
        name=request.form.get('name'),
        client_name=request.form.get('client_name'),
        address=request.form.get('address'),
        start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d') if request.form.get('start_date') else None,
        end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d') if request.form.get('end_date') else None,
        status=request.form.get('status'),
        description=request.form.get('description')
    )
    db.session.add(project)
    db.session.commit()
    flash('Project added successfully')
    return redirect(url_for('projects'))

@app.route('/project/<int:project_id>/contacts/add_new', methods=['POST'])
@login_required
def project_add_new_contact(project_id):
    project = Project.query.get_or_404(project_id)
    name = request.form.get('name')
    phone = request.form.get('phone')
    email = request.form.get('email')
    
    if name:
        # Create new contact
        contact = Contact(name=name, phone=phone, email=email, type='Client')
        db.session.add(contact)
        # Link to project
        project.contacts.append(contact)
        db.session.commit()
        flash('✅ Nouveau contact ajouté et lié au projet')
    return redirect(url_for('project_detail', id=project_id))

@app.route('/project/<int:project_id>/contacts/link_existing', methods=['POST'])
@login_required
def project_link_existing_contact(project_id):
    project = Project.query.get_or_404(project_id)
    contact_id = request.form.get('contact_id')
    if contact_id:
        contact = Contact.query.get(contact_id)
        if contact and contact not in project.contacts:
            project.contacts.append(contact)
            db.session.commit()
            flash('✅ Contact lié au projet')
    return redirect(url_for('project_detail', id=project_id))

@app.route('/project/<int:project_id>/contacts/<int:contact_id>/unlink', methods=['POST'])
@login_required
def project_unlink_contact(project_id, contact_id):
    project = Project.query.get_or_404(project_id)
    contact = Contact.query.get_or_404(contact_id)
    if contact in project.contacts:
        project.contacts.remove(contact)
        db.session.commit()
        flash('✅ Contact dissocié du projet')
    return redirect(url_for('project_detail', id=project_id))

@app.route('/projects/<int:id>')
@login_required
def project_detail(id):
    project = Project.query.get_or_404(id)
    all_contacts = Contact.query.all()
    return render_template('project_detail.html', project=project, all_contacts=all_contacts)


@app.route('/project/<int:id>/workers')
@login_required
def project_workers(id):
    """Render the standalone workers management page for a project."""
    project = Project.query.get_or_404(id)
    workers = ProjectWorker.query.filter_by(project_id=id).all()
    subcontractors = [w for w in workers if w.worker_type == 'subcontractor']
    daily_workers = [w for w in workers if w.worker_type == 'daily_worker']
    return render_template('project_workers.html', project=project, subcontractors=subcontractors, daily_workers=daily_workers)


@app.route('/project/<int:id>/workers-v2')
@login_required
def project_workers_v2(id):
    try:
        project = Project.query.get_or_404(id)
        print(f"DEBUG: Loading project {id}")
        
        subcontractors = ProjectWorker.query.filter_by(project_id=id, worker_type='subcontractor').options(db.joinedload(ProjectWorker.contact)).all()
        daily_workers = ProjectWorker.query.filter_by(project_id=id, worker_type='daily_worker').options(db.joinedload(ProjectWorker.contact)).all()
        contacts = Contact.query.all()
        
        print(f"DEBUG: Found {len(subcontractors)} subcontractors and {len(daily_workers)} daily workers")
        
        for w in daily_workers:
            print(f"DEBUG: Worker {w.id} - contact_id: {w.contact_id}, contact: {w.contact.name if w.contact else None}, days_worked: {w.days_worked}, daily_rate: {w.daily_rate}, amount_paid: {w.amount_paid}")
        
        print(f"DEBUG: Rendering template...")
        return render_template('project_workers_v2.html', project=project, subcontractors=subcontractors, daily_workers=daily_workers, contacts=contacts)
    except Exception as e:
        print(f"ERROR in project_workers_v2: {e}")
        import traceback
        traceback.print_exc()
        raise

@app.route('/projects/<int:id>', methods=['POST'])
@login_required
def update_project(id):
    project = Project.query.get_or_404(id)
    
    # Update basic fields
    project.name = request.form.get('name')
    project.client_name = request.form.get('client_name')
    project.status = request.form.get('status')
    project.address = request.form.get('address')
    project.description = request.form.get('description')
    
    # Handle dates
    start_date_str = request.form.get('start_date')
    if start_date_str:
        project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        project.start_date = None
        
    end_date_str = request.form.get('end_date')
    if end_date_str:
        project.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        project.end_date = None
    
    # Handle financial fields
    selling_price = request.form.get('selling_price', '0')
    project.selling_price = float(selling_price) if selling_price else 0.0
    
    client_receipts = request.form.get('client_receipts', '0')
    project.client_receipts = float(client_receipts) if client_receipts else 0.0
    
    db.session.commit()
    flash('Project updated successfully')
    
    return redirect(url_for('project_detail', id=id))

@app.route('/projects/<int:id>/update-status', methods=['POST'])
@login_required
def update_project_status(id):
    project = Project.query.get_or_404(id)
    data = request.get_json()
    
    if not data or 'status' not in data:
        return jsonify({'error': 'Status is required'}), 400
    
    new_status = data['status']
    if new_status not in ['planned', 'in_progress', 'completed']:
        return jsonify({'error': 'Invalid status'}), 400
    
    project.status = new_status
    db.session.commit()
    
    return jsonify({'success': True, 'status': project.status}), 200

@app.route('/projects/delete/<int:id>', methods=['POST'])
@login_required
@handle_errors
def delete_project(id):
    try:
        # First check if project exists
        project = Project.query.get(id)
        if not project:
            return jsonify({'error': 'Project not found. It may have been deleted already.'}), 404

        # Store project name for success message before deletion
        project_name = project.name

        # Delete related records first
        db.session.execute(project_contacts.delete().where(project_contacts.c.project_id == id))
        ProjectExpense.query.filter_by(project_id=id).delete(synchronize_session=False)
        ProjectReceipt.query.filter_by(project_id=id).delete(synchronize_session=False)
        ProjectFinancialParams.query.filter_by(project_id=id).delete(synchronize_session=False)
        ProjectMemo.query.filter_by(project_id=id).delete(synchronize_session=False)
        ProjectPlan.query.filter_by(project_id=id).delete(synchronize_session=False)

        # Check if Contract and Invoice models exist before trying to delete them
        try:
            Contract.query.filter_by(project_id=id).delete(synchronize_session=False)
            Invoice.query.filter_by(project_id=id).delete(synchronize_session=False)
        except Exception:
            # These models might not exist in the database schema, just continue
            app.logger.warning('Contract or Invoice model not found during deletion')

        # Handle workers and their related records
        worker_ids = [worker.id for worker in ProjectWorker.query.filter_by(project_id=id).all()]
        if worker_ids:
            Attendance.query.filter(Attendance.project_worker_id.in_(worker_ids)).delete(synchronize_session=False)
            Payment.query.filter(Payment.project_worker_id.in_(worker_ids)).delete(synchronize_session=False)
        ProjectWorker.query.filter_by(project_id=id).delete(synchronize_session=False)

        # Finally delete the project
        db.session.delete(project)
        db.session.commit()
        return jsonify({'success': True, 'message': f'Project "{project_name}" deleted successfully'})
    except Exception as e:
        db.session.rollback()
        app.logger.exception('Failed to delete project %s', id)
        return jsonify({'error': f'Failed to delete project: {str(e)}'}), 500

# ==================== TOOLS MANAGEMENT ROUTES ====================
@app.route('/budgets')
@login_required
def budgets():
    tools_list = Tool.query.all()
    return render_template('budgets.html', tools=tools_list)

@app.route('/budgets/add', methods=['POST'])
@login_required
def add_budget():
    from datetime import datetime
    tool = Tool(
        name=request.form.get('name'),
        category=request.form.get('category'),
        quantity=int(request.form.get('quantity', 1)),
        purchase_price=float(request.form.get('purchase_price')) if request.form.get('purchase_price') else None,
        current_value=float(request.form.get('current_value')) if request.form.get('current_value') else None,
        condition=request.form.get('condition', 'Good'),
        location=request.form.get('location'),
        purchase_date=datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d').date() if request.form.get('purchase_date') else None,
        notes=request.form.get('notes')
    )
    db.session.add(tool)
    db.session.commit()
    flash('Tool added successfully')
    return redirect(url_for('budgets'))

@app.route('/budgets/delete/<int:id>')
@login_required
def delete_budget(id):
    tool = Tool.query.get_or_404(id)
    db.session.delete(tool)
    db.session.commit()
    flash('Tool deleted')
    return redirect(url_for('budgets'))

# ==================== STUFF MANAGEMENT ROUTES ====================
@app.route('/stuff-management')
@login_required
def stuff_management():
    # Sample data for demonstration - replace with actual database queries when ready
    return render_template('stuff_management.html')

@app.route('/stuff/add', methods=['POST'])
@login_required
def add_stuff():
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'success': False,
                'message': 'No data received'
            }), 400

        # Validate required fields
        name = data.get('name')
        category = data.get('category')

        if not name or not category:
            return jsonify({
                'success': False,
                'message': 'Name and category are required'
            }), 400

        # For now, we'll store stuff as contacts with specific types
        # Later this should be moved to a proper Stuff model
        contact = Contact(
            name=name,
            phone=data.get('phone', ''),
            email=data.get('email', ''),
            type=f'Stuff-{category}',
            notes=f"""Category: {category}
Brand/Model: {data.get('brand', '')}
Serial: {data.get('serial', '')}
Quantity: {data.get('quantity', 1)}
Min Stock: {data.get('minStock', 1)}
Rate: {data.get('rate', 0)} DT/day
Price: {data.get('price', 0)} DT
Location: {data.get('location', '')}
Description: {data.get('description', '')}
Added on {datetime.now().strftime('%Y-%m-%d')}"""
        )

        db.session.add(contact)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{category.title()} item added successfully',
            'contact_id': contact.id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error adding stuff item: {str(e)}'
        }), 500

@app.route('/staff-management')
@login_required
def staff_management():
    # Sample data for demonstration - replace with actual database queries
    sample_staff = [
        {
            'id': 1,
            'name': 'Ahmed Ben Ali',
            'email': 'ahmed@company.com',
            'phone': '+216 22 123 456',
            'daily_rate': 80.0,
            'status': 'active',
            'hire_date': datetime(2024, 1, 15),
            'days_worked_this_month': 18,
            'current_projects': ['Villa Mahdia', 'Djerba Resort']
        },
        {
            'id': 2,
            'name': 'Mohamed Salem',
            'email': 'mohamed@company.com',
            'phone': '+216 98 765 432',
            'daily_rate': 75.0,
            'status': 'active',
            'hire_date': datetime(2024, 2, 1),
            'days_worked_this_month': 20,
            'current_projects': ['Villa Mahdia', 'Sousse Tower']
        },
        {
            'id': 3,
            'name': 'Karim Tounsi',
            'email': 'karim@company.com',
            'phone': '+216 55 444 333',
            'daily_rate': 70.0,
            'status': 'inactive',
            'hire_date': datetime(2024, 3, 10),
            'days_worked_this_month': 8,
            'current_projects': ['Djerba Resort']
        }
    ]
    
    return render_template('staff_management_demo.html', 
                         staff=sample_staff,
                         total_days_this_month=156,
                         monthly_payroll=12480,
                         current_month='March',
                         today_attendance=7,
                         pending_payments=3,
                         active_projects_with_staff=4)

@app.route('/workers/daily/create', methods=['POST'])
@login_required
def create_daily_worker():
    try:
        data = request.get_json()
        print(f"Received data: {data}")  # Debug log
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data received'
            }), 400
        
        # Validate required fields
        name = data.get('name')
        phone = data.get('phone')
        
        if not name or not phone:
            return jsonify({
                'success': False,
                'message': 'Name and phone are required'
            }), 400
        
        # Create contact first
        contact = Contact(
            name=name,
            phone=phone,
            email=data.get('email', ''),
            type='Daily Worker',
            notes=f"Rate: {data.get('rate', '')}/day\nCreated as daily worker on {datetime.now().strftime('%Y-%m-%d')}"
        )
        print(f"Created contact object: {contact}")  # Debug log
        
        db.session.add(contact)
        db.session.flush()  # Get the ID without committing
        print(f"Contact ID after flush: {contact.id}")  # Debug log
        
        # Here you would create the daily worker record
        # For now, we'll just return success since we don't have the Worker model yet
        db.session.commit()
        print("Database commit successful")  # Debug log
        
        return jsonify({
            'success': True,
            'message': 'Daily worker created successfully',
            'contact_id': contact.id
        })
        
    except Exception as e:
        print(f"Error in create_daily_worker: {str(e)}")  # Debug log
        import traceback
        traceback.print_exc()  # Print full traceback
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error creating daily worker: {str(e)}'
        }), 500

@app.route('/workers/daily/create-from-contact', methods=['POST'])
@login_required
def create_daily_worker_from_contact():
    try:
        data = request.get_json()
        contact_id = data.get('contact_id')
        
        if not contact_id:
            return jsonify({
                'success': False,
                'message': 'Contact ID is required'
            }), 400
        
        # Get the contact
        contact = Contact.query.get(contact_id)
        if not contact:
            return jsonify({
                'success': False,
                'message': 'Contact not found'
            }), 404
        
        # Update contact type to Daily Worker if not already set
        if contact.type != 'Daily Worker':
            contact.type = 'Daily Worker'
            contact.notes = (contact.notes or '') + f"\nConverted to daily worker on {datetime.now().strftime('%Y-%m-%d')}"
        
        # Here you would create the daily worker record
        # For now, we'll just update the contact
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Daily worker created from contact successfully',
            'contact_id': contact.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/workers/subcontractor/create', methods=['POST'])
@login_required
def create_subcontractor():
    try:
        data = request.get_json()
        
        # Create contact first
        contact = Contact(
            name=data.get('name'),
            phone=data.get('phone'),
            email=data.get('email'),
            type='Subcontractor',
            notes=f"Specialty: {data.get('specialty')}\nCreated as subcontractor on {datetime.now().strftime('%Y-%m-%d')}"
        )
        db.session.add(contact)
        db.session.flush()  # Get the ID without committing
        
        # Here you would create the subcontractor record
        # For now, we'll just return success since we don't have the Worker model yet
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subcontractor created successfully',
            'contact_id': contact.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/workers/subcontractor/create-from-contact', methods=['POST'])
@login_required
def create_subcontractor_from_contact():
    try:
        data = request.get_json()
        contact_id = data.get('contact_id')
        specialty = data.get('specialty')
        
        if not contact_id:
            return jsonify({
                'success': False,
                'message': 'Contact ID is required'
            }), 400
        
        if not specialty:
            return jsonify({
                'success': False,
                'message': 'Specialty is required'
            }), 400
        
        # Get the contact
        contact = Contact.query.get(contact_id)
        if not contact:
            return jsonify({
                'success': False,
                'message': 'Contact not found'
            }), 404
        
        # Update contact type to Subcontractor if not already set
        if contact.type != 'Subcontractor':
            contact.type = 'Subcontractor'
        
        # Update notes with specialty
        notes = (contact.notes or '') + f"\nSpecialty: {specialty}\nConverted to subcontractor on {datetime.now().strftime('%Y-%m-%d')}"
        contact.notes = notes.strip()
        
        # Here you would create the subcontractor record
        # For now, we'll just update the contact
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subcontractor created from contact successfully',
            'contact_id': contact.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# ==================== API ROUTES ====================
@app.route('/api/dashboard')
@login_required
def dashboard_data():
    total_contacts = Contact.query.count()
    total_providers = Provider.query.count()
    active_projects = Project.query.filter_by(status='in_progress').count()
    total_tools_value = db.session.query(db.func.sum(Tool.current_value)).scalar() or 0
    
    return jsonify({
        'contacts': total_contacts,
        'providers': total_providers,
        'active_projects': active_projects,
        'total_budget': total_tools_value
    })

@app.route('/api/projects/recent')
@login_required
def recent_projects():
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    query = Project.query
    
    if status and status != 'all':
        query = query.filter(Project.status == status)
    
    if search:
        query = query.filter(
            (Project.name.ilike(f'%{search}%')) | 
            (Project.client_name.ilike(f'%{search}%')) |
            (Project.address.ilike(f'%{search}%'))
        )
    
    projects = query.order_by(Project.id.desc()).all()
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'status': p.status,
        'client_name': p.client_name,
        'address': p.address,
        'start_date': p.start_date.strftime('%Y-%m-%d') if p.start_date else None,
        'end_date': p.end_date.strftime('%Y-%m-%d') if p.end_date else None,
        'selling_price': p.selling_price
    } for p in projects])

@app.route('/api/projects/list')
@login_required
def projects_list():
    try:
        projects = Project.query.order_by(Project.created_at.desc()).limit(10).all()
        return jsonify([{
            'id': p.id,
            'name': p.name,
            'status': p.status,
            'start_date': p.start_date.strftime('%Y-%m-%d') if p.start_date else None,
            'end_date': p.end_date.strftime('%Y-%m-%d') if p.end_date else None,
            'budget_total': p.budget_total,
            'address': p.address,
            'contact_name': p.client_name,
            'contact_phone': 'À venir'
        } for p in projects])
    except Exception as e:
        print(f"Error in projects_list: {e}")
        return jsonify([]), 500

# ==================== PROJECT WORKERS API ====================
@app.route('/api/project/<int:project_id>/workers')
@login_required
def get_project_workers(project_id):
    """Get all workers (subcontractors and daily workers) for a project"""
    try:
        workers = ProjectWorker.query.filter_by(project_id=project_id).all()
        return jsonify([{
            'id': w.id,
            'contact_id': w.contact_id,
            'name': w.contact.name if w.contact else 'Unknown',
            'phone': w.contact.phone if w.contact else '',
            'worker_type': w.worker_type,
            'role': w.role,
            'start_date': w.start_date.strftime('%Y-%m-%d') if w.start_date else None,
            'contract_amount': w.contract_amount,
            'daily_rate': w.daily_rate,
            'amount_paid': w.amount_paid,
            'remaining_balance': w.remaining_balance,
            'status': w.status,
            'notes': w.notes
        } for w in workers])
    except Exception as e:
        print(f"Error getting project workers: {e}")
        return jsonify([]), 500

@app.route('/api/project/<int:project_id>/workers', methods=['POST'])
@login_required
def add_project_worker(project_id):
    """Add a worker (subcontractor or daily worker) to a project"""
    try:
        data = request.get_json()
        # Accept either contact_id or contact_name. If name provided and not found, create a new Contact.
        contact = None
        if data.get('contact_id'):
            contact = Contact.query.get(data.get('contact_id'))
            if not contact:
                return jsonify({'success': False, 'error': 'Contact not found'}), 404
        else:
            name = (data.get('contact_name') or '').strip()
            if not name:
                return jsonify({'success': False, 'error': 'Missing contact identifier'}), 400
            # try to find existing contact by exact name
            contact = Contact.query.filter_by(name=name).first()
            if not contact:
                contact = Contact(name=name)
                db.session.add(contact)
                db.session.flush()  # assign id
        
        existing = ProjectWorker.query.filter_by(
            project_id=project_id,
            contact_id=contact.id
        ).first()
        
        if existing:
            return jsonify({'success': False, 'error': 'Contact already assigned to this project'}), 400
        
        worker = ProjectWorker(
            project_id=project_id,
            contact_id=contact.id,
            worker_type=data.get('worker_type'),
            role=data.get('role'),
            start_date=datetime.strptime(data.get('start_date'), '%Y-%m-%d').date() if data.get('start_date') else None,
            contract_amount=float(data.get('contract_amount', 0)),
            daily_rate=float(data.get('daily_rate', 0)),
            days_worked=float(data.get('days_worked', 0)),
            amount_paid=float(data.get('amount_paid', 0)) if data.get('amount_paid') is not None else 0,
            notes=data.get('notes')
        )
        
        db.session.add(worker)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'worker': {
                'id': worker.id,
                'contact_id': worker.contact_id,
                'name': worker.contact.name if worker.contact else 'Unknown',
                'worker_type': worker.worker_type,
                'role': worker.role,
                'days_worked': worker.days_worked
            }
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error adding project worker: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/workers/<int:worker_id>', methods=['POST', 'PUT'])
@login_required
def update_project_worker(project_id, worker_id):
    """Update a worker assignment for a project"""
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'}), 404

        data = request.get_json() or {}
        print(f"DEBUG: Updating worker {worker_id} with data: {data}")
        print(f"DEBUG: Current worker_type: {worker.worker_type}")
        # allow updating contact by id or by name (create contact if necessary)
        if 'contact_id' in data and data.get('contact_id'):
            contact = Contact.query.get(data.get('contact_id'))
            if contact:
                worker.contact_id = contact.id
        elif 'contact_name' in data and data.get('contact_name'):
            name = (data.get('contact_name') or '').strip()
            if name:
                contact = Contact.query.filter_by(name=name).first()
                if not contact:
                    contact = Contact(name=name)
                    db.session.add(contact)
                    db.session.flush()
                worker.contact_id = contact.id
        # Update allowed fields
        if 'role' in data:
            worker.role = data.get('role')
        if 'start_date' in data:
            sd = data.get('start_date')
            worker.start_date = datetime.strptime(sd, '%Y-%m-%d').date() if sd else None
        if 'contract_amount' in data:
            worker.contract_amount = float(data.get('contract_amount') or 0)
        if 'daily_rate' in data:
            worker.daily_rate = float(data.get('daily_rate') or 0)
        if 'days_worked' in data:
            worker.days_worked = float(data.get('days_worked') or 0)
        if 'amount_paid' in data:
            try:
                worker.amount_paid = float(data.get('amount_paid') or 0)
            except Exception:
                worker.amount_paid = 0
        if 'notes' in data:
            worker.notes = data.get('notes')
        # Preserve worker_type if provided
        if 'worker_type' in data:
            worker.worker_type = data.get('worker_type')

        db.session.commit()
        print(f"DEBUG: After commit - days_worked: {worker.days_worked}, daily_rate: {worker.daily_rate}, amount_paid: {worker.amount_paid}, worker_type: {worker.worker_type}")

        return jsonify({'success': True, 'worker': {
            'id': worker.id,
            'role': worker.role,
            'start_date': worker.start_date.strftime('%Y-%m-%d') if worker.start_date else None,
            'contract_amount': worker.contract_amount,
            'daily_rate': worker.daily_rate,
            'days_worked': worker.days_worked,
            'notes': worker.notes
        }})
    except Exception as e:
        db.session.rollback()
        print(f"Error updating worker: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/project/<int:project_id>/workers/<int:worker_id>', methods=['DELETE'])
@login_required
def remove_project_worker(project_id, worker_id):
    """Remove a worker from a project"""
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'}), 404
        
        db.session.delete(worker)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/project/<int:project_id>/workers/<int:worker_id>/payment', methods=['POST'])
@login_required
def record_worker_payment(project_id, worker_id):
    """Record a payment for a worker"""
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'}), 404
        
        data = request.get_json()
        amount = float(data.get('amount', 0))
        date_str = data.get('payment_date')
        method = data.get('payment_method')
        notes = data.get('notes')

        if amount <= 0:
            return jsonify({'success': False, 'error': 'Amount must be > 0'}), 400

        # parse date if provided
        if date_str:
            try:
                pay_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except Exception:
                pay_date = datetime.utcnow().date()
        else:
            pay_date = datetime.utcnow().date()

        payment = Payment(project_worker_id=worker.id, date=pay_date, amount=amount, method=method, notes=notes)
        db.session.add(payment)

        worker.amount_paid = (worker.amount_paid or 0) + amount
        db.session.commit()

        # return created payment info so frontend can use the real id
        return jsonify({
            'success': True,
            'amount_paid': worker.amount_paid,
            'remaining_balance': worker.remaining_balance,
            'payment': {
                'id': payment.id,
                'date': payment.date.strftime('%Y-%m-%d') if payment.date else None,
                'amount': payment.amount,
                'method': payment.method,
                'notes': payment.notes
            },
            'payment_id': payment.id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/workers/<int:worker_id>/payments')
@login_required
def get_worker_payments(project_id, worker_id):
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify([]), 404
        payments = Payment.query.filter_by(project_worker_id=worker.id).order_by(Payment.date.desc()).all()
        return jsonify([{
            'id': p.id,
            'date': p.date.strftime('%Y-%m-%d') if p.date else None,
            'amount': p.amount,
            'method': p.method,
            'notes': p.notes
        } for p in payments])
    except Exception as e:
        print(f"Error getting payments: {e}")
        return jsonify([]), 500


@app.route('/api/project/<int:project_id>/workers/<int:worker_id>/payments/<int:payment_id>', methods=['PUT'])
@login_required
def update_worker_payment(project_id, worker_id, payment_id):
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'}), 404
        payment = Payment.query.filter_by(id=payment_id, project_worker_id=worker.id).first()
        if not payment:
            return jsonify({'success': False, 'error': 'Payment not found'}), 404

        data = request.get_json() or {}
        old_amount = float(payment.amount or 0)
        new_amount = float(data.get('amount', old_amount))
        date_str = data.get('date')
        method = data.get('method') or data.get('payment_method') or payment.method
        notes = data.get('notes', payment.notes)

        if date_str:
            try:
                payment.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except Exception:
                pass

        payment.amount = new_amount
        payment.method = method
        payment.notes = notes

        # adjust worker total paid
        worker.amount_paid = (worker.amount_paid or 0) + (new_amount - old_amount)

        db.session.commit()
        return jsonify({'success': True, 'payment_id': payment.id, 'amount_paid': worker.amount_paid})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/workers/<int:worker_id>/payments/<int:payment_id>', methods=['DELETE'])
@login_required
def delete_worker_payment(project_id, worker_id, payment_id):
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'}), 404
        payment = Payment.query.filter_by(id=payment_id, project_worker_id=worker.id).first()
        if not payment:
            return jsonify({'success': False, 'error': 'Payment not found'}), 404

        amt = float(payment.amount or 0)
        db.session.delete(payment)
        worker.amount_paid = max(0, (worker.amount_paid or 0) - amt)
        db.session.commit()
        return jsonify({'success': True, 'amount_paid': worker.amount_paid})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/workers/<int:worker_id>/attendance')
@login_required
def get_worker_attendance(project_id, worker_id):
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify([]), 404
        records = Attendance.query.filter_by(project_worker_id=worker.id).order_by(Attendance.date.desc()).all()
        result = []
        for a in records:
            amount = 0
            try:
                amount = (a.days or 0) * (worker.daily_rate or 0)
            except Exception:
                amount = 0
            result.append({
                'id': a.id,
                'date': a.date.strftime('%Y-%m-%d') if a.date else None,
                'days': a.days,
                'amount': amount,
                'notes': a.notes if hasattr(a, 'notes') else None
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error getting attendance: {e}")
        return jsonify([]), 500


@app.route('/api/project/<int:project_id>/workers/<int:worker_id>/attendance', methods=['POST'])
@login_required
def record_worker_attendance(project_id, worker_id):
    """Record attendance (number of days) for a project worker."""
    try:
        worker = ProjectWorker.query.filter_by(id=worker_id, project_id=project_id).first()
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'}), 404

        data = request.get_json() or {}
        days = float(data.get('days', 0))
        date_str = data.get('date')
        if date_str:
            attend_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            attend_date = datetime.utcnow().date()

        if days <= 0:
            return jsonify({'success': False, 'error': 'Days must be > 0'}), 400

        notes = data.get('notes') if data else None
        attendance = Attendance(project_worker_id=worker.id, date=attend_date, days=days, notes=notes)
        db.session.add(attendance)
        db.session.commit()

        total_days = worker.calculated_days_worked
        return jsonify({'success': True, 'total_days': total_days})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== PROJECT TASKS (GANTT CHART) ====================
@app.route('/api/project/<int:project_id>/tasks')
@login_required
def get_project_tasks(project_id):
    """Get all tasks for a project for Gantt chart"""
    try:
        project = Project.query.get_or_404(project_id)
        tasks = ProjectTask.query.filter_by(project_id=project_id).all()
        
        result = []
        for task in tasks:
            result.append({
                'id': task.id,
                'name': task.name,
                'start_date': task.start_date.strftime('%Y-%m-%d') if task.start_date else None,
                'end_date': task.end_date.strftime('%Y-%m-%d') if task.end_date else None,
                'progress': task.progress,
                'parent': task.parent_id,
                'type': 'task',
                'assigned_to': task.assigned_to,
                'status': task.status,
                'description': task.description
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/tasks', methods=['POST'])
@login_required
def save_project_tasks(project_id):
    """Save tasks (create, update, delete) for a project"""
    try:
        project = Project.query.get_or_404(project_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        # Handle single task save
        if 'action' in data and data['action'] == 'save':
            task_data = data.get('task', {})
            task_id = task_data.get('id')
            
            if task_id:
                # Update existing task
                task = ProjectTask.query.filter_by(id=task_id, project_id=project_id).first()
                if not task:
                    return jsonify({'success': False, 'error': 'Task not found'}), 404
                
                task.name = task_data.get('name', task.name)
                task.description = task_data.get('description', task.description)
                if task_data.get('start_date'):
                    task.start_date = datetime.strptime(task_data['start_date'], '%Y-%m-%d').date()
                if task_data.get('end_date'):
                    task.end_date = datetime.strptime(task_data['end_date'], '%Y-%m-%d').date()
                task.progress = task_data.get('progress', task.progress)
                task.parent_id = task_data.get('parent_id', task.parent_id)
                task.assigned_to = task_data.get('assigned_to', task.assigned_to)
                task.updated_at = datetime.utcnow()
            else:
                # Create new task
                task = ProjectTask(
                    project_id=project_id,
                    name=task_data.get('name'),
                    description=task_data.get('description'),
                    start_date=datetime.strptime(task_data['start_date'], '%Y-%m-%d').date() if task_data.get('start_date') else None,
                    end_date=datetime.strptime(task_data['end_date'], '%Y-%m-%d').date() if task_data.get('end_date') else None,
                    progress=task_data.get('progress', 0),
                    parent_id=task_data.get('parent_id'),
                    assigned_to=task_data.get('assigned_to')
                )
                db.session.add(task)
            
            db.session.commit()
            return jsonify({'success': True, 'task_id': task.id})
        
        # Handle delete action
        elif 'action' in data and data['action'] == 'delete':
            task_id = data.get('id')
            if not task_id:
                return jsonify({'success': False, 'error': 'Task ID required'}), 400
            
            task = ProjectTask.query.filter_by(id=task_id, project_id=project_id).first()
            if not task:
                return jsonify({'success': False, 'error': 'Task not found'}), 404
            
            db.session.delete(task)
            db.session.commit()
            return jsonify({'success': True})
        
        # Handle batch operations
        elif 'tasks' in data:
            for task_data in data['tasks']:
                action = task_data.get('action')
                
                if action == 'create':
                    task = ProjectTask(
                        project_id=project_id,
                        name=task_data.get('name'),
                        start_date=datetime.strptime(task_data['start_date'], '%Y-%m-%d').date() if task_data.get('start_date') else None,
                        end_date=datetime.strptime(task_data['end_date'], '%Y-%m-%d').date() if task_data.get('end_date') else None,
                        progress=task_data.get('progress', 0)
                    )
                    db.session.add(task)
                
                elif action == 'update':
                    task = ProjectTask.query.filter_by(id=task_data['id'], project_id=project_id).first()
                    if task:
                        if 'progress' in task_data:
                            task.progress = task_data['progress']
                        if 'end_date' in task_data:
                            task.end_date = datetime.strptime(task_data['end_date'], '%Y-%m-%d').date()
                        task.updated_at = datetime.utcnow()
                
                elif action == 'delete':
                    task = ProjectTask.query.filter_by(id=task_data['id'], project_id=project_id).first()
                    if task:
                        db.session.delete(task)
            
            db.session.commit()
            return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'Invalid action'}), 400
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/tasks/all', methods=['DELETE'])
@login_required
def delete_all_project_tasks(project_id):
    """Delete all tasks for a project"""
    try:
        project = Project.query.get_or_404(project_id)

        # Delete all tasks for the project
        ProjectTask.query.filter_by(project_id=project_id).delete()
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/tasks/dependencies')
@login_required
def get_task_dependencies(project_id):
    """Get task dependencies for a project"""
    try:
        project = Project.query.get_or_404(project_id)
        tasks = ProjectTask.query.filter_by(project_id=project_id).all()
        
        result = []
        for task in tasks:
            if task.dependency_type and task.parent_id:
                result.append({
                    'id': task.id,
                    'source': task.parent_id,
                    'target': task.id,
                    'type': task.dependency_type
                })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/project/<int:project_id>/tasks/dependencies', methods=['POST'])
@login_required
def save_task_dependencies(project_id):
    """Save task dependencies for a project"""
    try:
        project = Project.query.get_or_404(project_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        action = data.get('action')
        
        if action == 'save':
            link = data.get('link', {})
            task_id = link.get('target')
            
            task = ProjectTask.query.filter_by(id=task_id, project_id=project_id).first()
            if not task:
                return jsonify({'success': False, 'error': 'Task not found'}), 404
            
            task.parent_id = link.get('source')
            task.dependency_type = link.get('type')
            task.updated_at = datetime.utcnow()
            
            db.session.commit()
            return jsonify({'success': True})
        
        elif action == 'delete':
            link_id = data.get('id')
            
            task = ProjectTask.query.filter_by(id=link_id, project_id=project_id).first()
            if not task:
                return jsonify({'success': False, 'error': 'Task not found'}), 404
            
            task.parent_id = None
            task.dependency_type = None
            task.updated_at = datetime.utcnow()
            
            db.session.commit()
            return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'Invalid action'}), 400
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== CONTACT UPLOAD ====================
@app.route('/contacts/upload', methods=['POST'])
@login_required
def upload_contacts():
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('contacts'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('contacts'))
    
    try:
        if file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            rows = read_spreadsheet(file)
        else:
            flash('Please upload CSV or Excel file')
            return redirect(url_for('contacts'))

        success_count = 0
        error_count = 0
        errors = []

        for index, row in enumerate(rows):
            try:
                contact = Contact(
                    name=clean(row.get('Name', '')),
                    phone=clean(row.get('Phone', '')),
                    email=clean(row.get('Email', '')),
                    company=clean(row.get('Company', '')),
                    type='client',
                    notes=f"Speciality: {clean(row.get('Speciality', ''))} | Address: {clean(row.get('Address', ''))} | Comments: {clean(row.get('Comments', ''))}"
                )

                if not contact.name:
                    errors.append(f"Row {index + 2}: Name is required")
                    error_count += 1
                    continue

                db.session.add(contact)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                error_count += 1
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'✅ Successfully imported {success_count} contacts')
        if error_count > 0:
            flash(f'⚠️ Failed to import {error_count} contacts.')
            for error in errors[:5]:
                flash(f'Error: {error}')
        
    except Exception as e:
        flash(f'Error reading file: {str(e)}')
    
    return redirect(url_for('contacts'))

# ==================== PROVIDER UPLOAD ====================
@app.route('/providers/upload', methods=['POST'])
@login_required
def upload_providers():
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('providers'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('providers'))
    
    try:
        if file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            rows = read_spreadsheet(file)
        else:
            flash('Please upload CSV or Excel file')
            return redirect(url_for('providers'))

        success_count = 0
        error_count = 0
        errors = []

        for index, row in enumerate(rows):
            try:
                provider = Provider(
                    name=clean(row.get('Company Name', '')),
                    contact_person=clean(row.get('Contact Person', '')),
                    phone=clean(row.get('Phone', '')),
                    email=clean(row.get('Email', '')),
                    address=clean(row.get('Address', '')),
                    service_type=clean(row.get('Speciality', '')),
                    notes=clean(row.get('Comments', ''))
                )
                
                if not provider.name:
                    errors.append(f"Row {index + 2}: Company Name is required")
                    error_count += 1
                    continue
                
                db.session.add(provider)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                error_count += 1
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'✅ Successfully imported {success_count} providers')
        if error_count > 0:
            flash(f'⚠️ Failed to import {error_count} providers.')
            for error in errors[:5]:
                flash(f'Error: {error}')
        
    except Exception as e:
        flash(f'Error reading file: {str(e)}')
    
    return redirect(url_for('providers'))

# ==================== DOWNLOAD TEMPLATES ====================
@app.route('/contacts/template')
@login_required
def download_contacts_template():
    data = [
        {'Name': 'Ahmed Ben Ali', 'Phone': '+216 22 123 456', 'Email': 'ahmed@email.com', 'Company': 'ABC Construction', 'Speciality': 'Project Management', 'Comments': 'Good client, multiple projects', 'Address': 'Tunis, Centre Ville'},
        {'Name': 'Sarra Mansour', 'Phone': '+216 55 789 012', 'Email': 'sarra@email.com', 'Company': 'XYZ Materials', 'Speciality': 'Electrical Engineering', 'Comments': 'Reliable supplier', 'Address': 'Sousse, Rue de la Liberté'}
    ]
    output = write_to_excel_buffer(data, sheet_name='Contacts')

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='contacts_template.xlsx'
    )

@app.route('/providers/template')
@login_required
def download_providers_template():
    data = [
        {'Company Name': 'Matériaux Tunisie', 'Contact Person': 'Karim Ben Salem', 'Phone': '+216 71 123 456', 'Email': 'karim@materiaux.tn', 'Speciality': 'Construction Materials', 'Comments': 'Good prices, fast delivery', 'Address': 'Tunis, Zone Industrielle'},
        {'Company Name': 'Électro Plus', 'Contact Person': 'Leila Mansour', 'Phone': '+216 72 789 012', 'Email': 'leila@electroplus.tn', 'Speciality': 'Electrical Supplies', 'Comments': 'Certified products', 'Address': 'Sousse, Route de la Plage'}
    ]
    output = write_to_excel_buffer(data, sheet_name='Providers')

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='providers_template.xlsx'
    )

@app.route('/financial/template')
@login_required
def download_financial_template():
    """Download the financial Excel template"""
    try:
        # Path to your actual template file
        template_path = os.path.join(app.root_path, 'templates', 'financial_template.xlsx')
        
        # Check if file exists
        if not os.path.exists(template_path):
            flash('Template file not found. Please contact administrator.')
            return redirect(request.referrer or url_for('index'))
        
        return send_file(
            template_path,
            as_attachment=True,
            download_name='financial_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error downloading template: {str(e)}')
        return redirect(request.referrer or url_for('index'))

# ==================== DOWNLOAD CONTACTS ====================
@app.route('/contacts/download-all')
@login_required
def download_all_contacts():
    try:
        contacts = Contact.query.all()
        
        data = []
        for contact in contacts:
            notes = contact.notes or ""
            data.append({
                'Name': contact.name,
                'Email': contact.email or '',
                'Phone': contact.phone or '',
                'Company': contact.company or '',
                'Type': contact.type or '',
                'Notes': notes
            })

        output = write_to_excel_buffer(data, sheet_name='All Contacts')
        
        from datetime import datetime
        filename = f"all_contacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'Error downloading contacts: {str(e)}')
        return redirect(url_for('contacts'))
    
@app.route('/contacts/download-selected', methods=['POST'])
@login_required
def download_selected_contacts():
    try:
        selected_ids = json.loads(request.form.get('selected_ids', '[]'))
        
        if not selected_ids:
            flash('No contacts selected')
            return redirect(url_for('contacts'))
        
        contacts = Contact.query.filter(Contact.id.in_(selected_ids)).all()
        
        data = []
        for contact in contacts:
            data.append({
                'Name': contact.name,
                'Email': contact.email or '',
                'Phone': contact.phone or '',
                'Company': contact.company or '',
                'Type': contact.type or '',
                'Notes': contact.notes or ''
            })

        output = write_to_excel_buffer(data, sheet_name='Selected Contacts')
        
        from datetime import datetime
        filename = f"selected_contacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'Error downloading contacts: {str(e)}')
        return redirect(url_for('contacts'))

# ==================== PROJECT FINANCIAL ROUTES ====================
@app.route('/project/<int:project_id>/financial')
@login_required
def project_financial(project_id):
    project = Project.query.get_or_404(project_id)
    
    financial = ProjectFinancialParams.query.filter_by(project_id=project_id).first()
    if not financial:
        financial = ProjectFinancialParams(project_id=project_id)
        db.session.add(financial)
        db.session.commit()
    
    expenses = ProjectExpense.query.filter_by(project_id=project_id).order_by(ProjectExpense.date.desc()).all()
    total_expenses = sum(e.amount for e in expenses)
    
    receipts = ProjectReceipt.query.filter_by(project_id=project_id).all()
    # Use project.selling_price and project.client_receipts (same as project detail page)
    sale_price = project.selling_price or financial.sale_price or 0
    client_received = project.client_receipts if project.client_receipts is not None else 0
    current_cash = client_received - total_expenses
    estimated_margin = sale_price - financial.estimated_budget
    current_result = sale_price - total_expenses
    balance_remaining = sale_price - client_received  # Solde Restant (amount client still owes)
    
    categories = {}
    for expense in expenses:
        categories[expense.category] = categories.get(expense.category, 0) + expense.amount
    
    category_totals = []
    colors = {'Ouvrier': '#4DA8DA', 'Matériau': '#28a745', 'Divers': '#FF6600'}  # accent-glow, success, accent
    for cat, amount in categories.items():
        category_totals.append({
            'category': cat,
            'total': amount,
            'percentage': (amount / total_expenses * 100) if total_expenses > 0 else 0,
            'color': colors.get(cat, '#6c757d')
        })
    
    return render_template('project_financial.html', 
                         project=project,
                         financial={
                             'client_received': client_received,
                             'current_cash': current_cash,
                             'sale_price': sale_price,
                             'estimated_budget': financial.estimated_budget,
                             'estimated_margin': estimated_margin,
                             'total_expenses': total_expenses,
                             'current_result': current_result,
                             'depenses_reelles': total_expenses, # données projet #5: sum of expenses
                             'gain_actuel': current_result,      # données projet #6: Prix de Vente - Dépenses Réelles
                             'devis_ref': financial.devis_ref,
                             'balance_remaining': balance_remaining
                         },
                         expenses=expenses,
                         receipts=receipts,
                         total_expenses=total_expenses,
                         category_totals=category_totals)

@app.route('/project/<int:project_id>/expense/add', methods=['POST'])
@login_required
def add_project_expense(project_id):
    expense = ProjectExpense(
        project_id=project_id,
        date=datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
        nature=request.form.get('nature'),
        amount=float(request.form.get('amount')),
        comment=request.form.get('comment'),
        category=request.form.get('category')
    )
    db.session.add(expense)
    db.session.commit()
    flash('Expense added successfully')
    return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/receipt/add', methods=['POST'])
@login_required
def add_client_receipt(project_id):
    receipt = ProjectReceipt(
        project_id=project_id,
        date=datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
        amount=float(request.form.get('amount')),
        comment=request.form.get('comment')
    )
    db.session.add(receipt)
    db.session.commit()
    flash('Client receipt added successfully')
    return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/financial/update', methods=['POST'])
@login_required
def update_project_financial_params(project_id):
    financial = ProjectFinancialParams.query.filter_by(project_id=project_id).first()
    if not financial:
        financial = ProjectFinancialParams(project_id=project_id)
        db.session.add(financial)
    
    financial.sale_price = float(request.form.get('sale_price', 0))
    financial.estimated_budget = float(request.form.get('estimated_budget', 0))
    financial.devis_ref = float(request.form.get('devis_ref', 0))
    
    db.session.commit()
    flash('Project parameters updated')
    return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/expense/<int:expense_id>/edit', methods=['POST'])
@login_required
def edit_expense(project_id, expense_id):
    expense = ProjectExpense.query.get_or_404(expense_id)
    if expense.project_id != project_id:
        return redirect(url_for('project_financial', project_id=project_id))
    expense.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
    expense.nature = request.form.get('nature')
    expense.amount = float(request.form.get('amount'))
    expense.comment = request.form.get('comment') or None
    expense.category = request.form.get('category', 'Divers')
    db.session.commit()
    flash('Expense updated successfully')
    return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/expense/<int:expense_id>/delete', methods=['POST'])
@login_required
def delete_expense(project_id, expense_id):
    expense = ProjectExpense.query.get_or_404(expense_id)
    db.session.delete(expense)
    db.session.commit()
    flash('Expense deleted successfully')
    return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/financial/delete-all', methods=['POST'])
@login_required
def delete_all_project_financial(project_id):
    """Delete all financial data for a project (expenses + receipts)"""
    try:
        ProjectExpense.query.filter_by(project_id=project_id).delete()
        ProjectReceipt.query.filter_by(project_id=project_id).delete()
        db.session.commit()
        flash('✅ All financial data deleted successfully')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting data: {str(e)}')
    return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/expenses/delete-all', methods=['POST'])
@login_required
def delete_all_expenses(project_id):
    """Delete all expenses for a project (expenses table only, not receipts)"""
    try:
        deleted = ProjectExpense.query.filter_by(project_id=project_id).delete()
        db.session.commit()
        flash(f'✅ {deleted} dépenses supprimées')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erreur: {str(e)}')
    return redirect(url_for('project_financial', project_id=project_id))

# ==================== PROJECT FINANCIAL EXPORT/IMPORT ====================
@app.route('/project/<int:project_id>/financial/export')
@login_required
def export_project_financial(project_id):
    """Export project financial data to Excel"""
    try:
        project = Project.query.get_or_404(project_id)
        expenses = ProjectExpense.query.filter_by(project_id=project_id).all()
        
        # Build data from expenses
        data = []
        for expense in expenses:
            data.append({
                'Date': expense.date.strftime('%d/%m/%Y'),
                'Nature': expense.nature,
                'Montant': expense.amount,
                'Commentaire': expense.comment or '',
                'Catégorie': expense.category
            })

        output = write_to_excel_buffer(data, sheet_name='Dépenses')
        
        filename = f"expenses_{project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Error exporting: {str(e)}')
        return redirect(url_for('project_financial', project_id=project_id))

@app.route('/project/<int:project_id>/financial/import', methods=['POST'])
@login_required
def import_project_financial(project_id):
    """Import expenses from Excel - follows financial_template.xlsx (sheet 'suivi chantier').
    Imported rows are ADDED to expenses, not replaced."""
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('project_financial', project_id=project_id))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('project_financial', project_id=project_id))
    
    try:
        # Read the file as rows of dict
        if file.filename.lower().endswith('.csv'):
            rows = read_spreadsheet(file)
        else:
            # Excel file may have a specific sheet
            try:
                rows = read_spreadsheet(file, sheet_name='suivi chantier')
            except Exception:
                rows = read_spreadsheet(file)

        # Track success/failure
        success_count = 0
        error_count = 0
        errors = []

        def clean(val):
            if is_na(val):
                return ''
            return str(val).strip()

        # Process each row - ADD to expenses (never replace)
        for index, row in enumerate(rows):
            try:
                # Parse date (Excel may return datetime, or string)
                val = row.get('Date')
                if is_na(val):
                    date = datetime.now().date()
                elif hasattr(val, 'date'):
                    date = val.date()
                else:
                    date_str = clean(val)
                    if date_str:
                        try:
                            date = datetime.strptime(date_str, '%d/%m/%Y').date()
                        except:
                            try:
                                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                            except:
                                date = datetime.now().date()
                    else:
                        date = datetime.now().date()

                amount_str = clean(row.get('Montant', '0'))
                try:
                    amount = float(str(amount_str).replace(',', '.'))
                except:
                    amount = 0.0
                
                cat_val = 'Divers'
                for c in df.columns:
                    if 'cat' in c.lower():
                        cv = clean(row.get(c, 'Divers'))
                        if cv in ('Ouvrier', 'Matériau', 'Divers'):
                            cat_val = cv
                        elif 'ouvrier' in cv.lower():
                            cat_val = 'Ouvrier'
                        elif 'materiau' in cv.lower() or 'matériau' in cv.lower():
                            cat_val = 'Matériau'
                        break
                
                expense = ProjectExpense(
                    project_id=project_id,
                    date=date,
                    nature=clean(row.get('Nature', '')),
                    amount=amount,
                    comment=clean(row.get('Commentaire', '')),
                    category=cat_val
                )
                
                # Basic validation
                if not expense.nature:
                    errors.append(f"Row {index + 2}: Nature is required")
                    error_count += 1
                    continue
                
                if expense.amount <= 0:
                    errors.append(f"Row {index + 2}: Amount must be positive")
                    error_count += 1
                    continue
                
                db.session.add(expense)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                error_count += 1
        
        # Commit all successful entries
        db.session.commit()
        
        # Flash summary
        if success_count > 0:
            flash(f'✅ Successfully imported {success_count} expenses')
        if error_count > 0:
            flash(f'⚠️ Failed to import {error_count} expenses. Check your file format.')
            for error in errors[:5]:
                flash(f'Error: {error}')
        
    except Exception as e:
        flash(f'Error reading file: {str(e)}')
    
    return redirect(url_for('project_financial', project_id=project_id))

# ==================== DOCUMENT MANAGEMENT ROUTES ====================
@app.route('/project/<int:project_id>/memos')
@login_required
def project_memos(project_id):
    """View all memos for a project"""
    project = Project.query.get_or_404(project_id)
    memos = ProjectMemo.query.filter_by(project_id=project_id).order_by(ProjectMemo.date.desc()).all()
    return render_template('project_docs/memos.html', project=project, memos=memos)

@app.route('/project/<int:project_id>/memos/add', methods=['POST'])
@login_required
def add_project_memo(project_id):
    """Add a new memo to a project"""
    project = Project.query.get_or_404(project_id)
    
    memo = ProjectMemo(
        project_id=project_id,
        title=request.form.get('memo_title'),
        content=request.form.get('memo_content'),
        date=datetime.strptime(request.form.get('memo_date'), '%Y-%m-%d').date(),
        created_at=datetime.utcnow()
    )
    
    db.session.add(memo)
    db.session.commit()
    flash('✅ Mémo ajouté avec succès!')
    return redirect(url_for('project_memos', project_id=project_id))

@app.route('/project/<int:project_id>/memos/<int:memo_id>/delete', methods=['POST'])
@login_required
def delete_project_memo(project_id, memo_id):
    """Delete a memo"""
    memo = ProjectMemo.query.get_or_404(memo_id)
    if memo.project_id != project_id:
        flash('❌ Accès non autorisé')
        return redirect(url_for('index'))
    
    db.session.delete(memo)
    db.session.commit()
    flash('✅ Mémo supprimé avec succès!')
    return redirect(url_for('project_memos', project_id=project_id))

@app.route('/project/<int:project_id>/plans')
@login_required
def project_plans(project_id):
    """View all plans for a project"""
    project = Project.query.get_or_404(project_id)
    plans = ProjectPlan.query.filter_by(project_id=project_id).order_by(ProjectPlan.date.desc()).all()
    return render_template('project_docs/plans.html', project=project, plans=plans)

@app.route('/project/<int:project_id>/plans/add', methods=['POST'])
@login_required
def add_project_plan(project_id):
    """Add a new plan to a project"""
    project = Project.query.get_or_404(project_id)
    
    plan = ProjectPlan(
        project_id=project_id,
        title=request.form.get('title'),
        description=request.form.get('description'),
        date=datetime.strptime(request.form.get('date'), '%Y-%m-%d').date(),
        created_at=datetime.utcnow()
    )
    
    # Handle file upload
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename:
            filename = secure_filename(file.filename)
            file_path = os.path.join('uploads', 'plans', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            plan.file_path = file_path
    
    db.session.add(plan)
    db.session.commit()
    flash('✅ Plan ajouté avec succès!')
    return redirect(url_for('project_plans', project_id=project_id))

@app.route('/project/<int:project_id>/plans/<int:plan_id>/delete', methods=['POST'])
@login_required
def delete_project_plan(project_id, plan_id):
    """Delete a plan"""
    plan = ProjectPlan.query.get_or_404(plan_id)
    if plan.project_id != project_id:
        flash('❌ Accès non autorisé')
        return redirect(url_for('index'))
    
    # Optional: Delete physical file
    if plan.file_path and os.path.exists(plan.file_path):
        try:
            os.remove(plan.file_path)
        except Exception as e:
            print(f"Error deleting file: {e}")
            
    db.session.delete(plan)
    db.session.commit()
    flash('✅ Plan supprimé avec succès!')
    return redirect(url_for('project_plans', project_id=project_id))

@app.route('/project/<int:project_id>/contracts')
@login_required
def project_contracts(project_id):
    """View all contracts for a project"""
    project = Project.query.get_or_404(project_id)
    contracts = ProjectContract.query.filter_by(project_id=project_id).order_by(ProjectContract.date.desc()).all()
    return render_template('project_docs/contracts.html', project=project, contracts=contracts)

@app.route('/project/<int:project_id>/contracts/add', methods=['POST'])
@login_required
def add_project_contract(project_id):
    """Add a new contract to a project"""
    project = Project.query.get_or_404(project_id)
    
    contract = ProjectContract(
        project_id=project_id,
        title=request.form.get('title'),
        description=request.form.get('description'),
        date=datetime.strptime(request.form.get('date'), '%Y-%m-%d').date(),
        created_at=datetime.utcnow()
    )
    
    # Handle file upload
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename:
            filename = secure_filename(file.filename)
            file_path = os.path.join('uploads', 'contracts', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            contract.file_path = file_path
    
    db.session.add(contract)
    db.session.commit()
    flash('✅ Contrat ajouté avec succès!')
    return redirect(url_for('project_contracts', project_id=project_id))

@app.route('/project/<int:project_id>/contracts/<int:contract_id>/delete', methods=['POST'])
@login_required
def delete_project_contract(project_id, contract_id):
    """Delete a contract"""
    contract = ProjectContract.query.get_or_404(contract_id)
    if contract.project_id != project_id:
        flash('❌ Accès non autorisé')
        return redirect(url_for('index'))
    
    if contract.file_path and os.path.exists(contract.file_path):
        try:
            os.remove(contract.file_path)
        except Exception as e:
            print(f"Error deleting file: {e}")
            
    db.session.delete(contract)
    db.session.commit()
    flash('✅ Contrat supprimé avec succès!')
    return redirect(url_for('project_contracts', project_id=project_id))

@app.route('/project/<int:project_id>/invoices')
@login_required
def project_invoices(project_id):
    """View all invoices for a project"""
    project = Project.query.get_or_404(project_id)
    invoices = ProjectInvoice.query.filter_by(project_id=project_id).order_by(ProjectInvoice.date.desc()).all()
    return render_template('project_docs/invoices.html', project=project, invoices=invoices)

@app.route('/project/<int:project_id>/invoices/add', methods=['POST'])
@login_required
def add_project_invoice(project_id):
    """Add a new invoice to a project"""
    project = Project.query.get_or_404(project_id)
    
    invoice = ProjectInvoice(
        project_id=project_id,
        title=request.form.get('title'),
        description=request.form.get('description'),
        date=datetime.strptime(request.form.get('date'), '%Y-%m-%d').date(),
        created_at=datetime.utcnow()
    )
    
    # Handle file upload
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename:
            filename = secure_filename(file.filename)
            file_path = os.path.join('uploads', 'invoices', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            invoice.file_path = file_path
    
    db.session.add(invoice)
    db.session.commit()
    flash('✅ Facture ajoutée avec succès!')
    return redirect(url_for('project_invoices', project_id=project_id))

@app.route('/project/<int:project_id>/invoices/<int:invoice_id>/delete', methods=['POST'])
@login_required
def delete_project_invoice(project_id, invoice_id):
    """Delete an invoice"""
    invoice = ProjectInvoice.query.get_or_404(invoice_id)
    if invoice.project_id != project_id:
        flash('❌ Accès non autorisé')
        return redirect(url_for('index'))
    
    if invoice.file_path and os.path.exists(invoice.file_path):
        try:
            os.remove(invoice.file_path)
        except Exception as e:
            print(f"Error deleting file: {e}")
            
    db.session.delete(invoice)
    db.session.commit()
    flash('✅ Facture supprimée avec succès!')
    return redirect(url_for('project_invoices', project_id=project_id))

@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    # Check if download parameter is in request
    if request.args.get('download'):
        return send_from_directory('uploads', filename, as_attachment=True)
    return send_from_directory('uploads', filename)

# ==================== MAIN (entry point is at bottom) ====================
# Ensure runtime migrations run when using flask's server (flask run)
def ensure_runtime_migrations():
    try:
        with app.app_context():
            db.create_all()
            res = db.session.execute(text("PRAGMA table_info('attendance')")).fetchall()
            cols = [row[1] for row in res]
            if 'notes' not in cols:
                try:
                    db.session.execute(text("ALTER TABLE attendance ADD COLUMN notes TEXT"))
                    db.session.commit()
                    print("Added 'notes' column to attendance table (before_first_request)")
                except Exception as inner_e:
                    db.session.rollback()
                    print('Failed to add notes column to attendance (before_first_request):', inner_e)

            # Add days_worked column to project_worker table if it doesn't exist
            res = db.session.execute(text("PRAGMA table_info('project_worker')")).fetchall()
            cols = [row[1] for row in res]
            if 'days_worked' not in cols:
                try:
                    db.session.execute(text("ALTER TABLE project_worker ADD COLUMN days_worked FLOAT DEFAULT 0.0"))
                    db.session.commit()
                    print("Added 'days_worked' column to project_worker table")
                except Exception as inner_e:
                    db.session.rollback()
                    print('Failed to add days_worked column to project_worker:', inner_e)
    except Exception as e:
        print('ensure_runtime_migrations failed:', e)

# Register migration handler if supported (some Flask runtimes expose before_first_request)
if hasattr(app, 'before_first_request'):
    app.before_first_request(ensure_runtime_migrations)

if __name__ == '__main__':
    # When running directly, run migrations and start server
    with app.app_context():
        ensure_runtime_migrations()
    app.run(debug=True, host='0.0.0.0', port=5000)